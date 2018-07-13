#!/usr/local/bin/python3
"""
--------- To Do --------
make threads start from the top after successfully looping in a task
simultaneous search requests for the same query with different page ranges (?)

------- Samples ----------
r = requests.get("https://flexshopper.com/search-chunk?termOrSlug=camera&range%5B%5D=1&range%5B%5D=83")

"""
import openpyxl
import requests
import bs4
import threading
import queue
from tkinter import *
from tkinter import filedialog # don't know why this doesn't come with above command
from platform import system
import time
import datetime
from re import findall

root = Tk()
root.title("Find Flexshopper Products That Are Missing Images")
root.geometry('{}x{}'.format(700, 600))
root.minsize(600,550)
# root.configure(background="red")
root["bg"] = "#e6e6e6"

if system() == "Darwin":
    fontsetting = ("Arial",13)
    # root.wm_iconbitmap('fslogo.icns')
    print("System is Darwin")
    
if system() == "Windows":
    fontsetting = ("Arial",10)
    root.wm_iconbitmap('fslogo.ico')
    print("System is Windows")

print(system())

updatesQueue = queue.Queue() # Updates for the GUI

queryQueue = queue.Queue()
pageQueue = queue.Queue()
DomQueue = queue.Queue()
needIdQueue = queue.Queue()

queryThreads = []
pageThreads = []
DomThreads = []
needIdThreads = []


earlyAbort = threading.Event()
maxThreads = 100
searchCriteria = StringVar()
consecutiveErrors = 0
duplicateCount = 0

def makeExcelDoc():
    global emptyListWindow
    if not "productList" in globals().keys() or len(productList) == 0: # Make sure there's a product list with stuff in it
        if not "emptyListWindow" in globals().keys() or emptyListWindow.winfo_exists() == 0: # if there isn't already an error message up
            emptyListWindow = Toplevel(bg = "#e6e6e6")
            emptyListWindow.title("No Products To List")
            # emptyListWindow.wm_iconbitmap('fslogo.ico')
            if system() == "Darwin":
                pass
                # emptyListWindow.wm_iconbitmap('fslogo.icns')
            if system() == "Windows":
                emptyListWindow.wm_iconbitmap('fslogo.ico')
            errorText = "List of missing image products is empty.\nRun a search with results first."
            emptyMessage = Label(emptyListWindow, text=errorText, bg = "#e6e6e6", name="emptyMessage")
            errorOkButton = Button(emptyListWindow, text="OK", highlightbackground = "#e6e6e6", name="errorOkButton", command=emptyListWindow.destroy)
            emptyMessage.pack(side=TOP, anchor=N, padx=40, pady=(20,10))
            errorOkButton.pack(side=TOP, anchor=S, padx=40, pady=(0,20))
        else: #                                                              # if there IS already an error message up
            emptyListWindow.attributes('-topmost', 1) # Raise that window to the top temporarily
            emptyListWindow.attributes('-topmost', 0)
            def changeColor(): # make the window blink
                currentColor = emptyListWindow.cget("bg")
                nextColor = "#ffffff" if currentColor == "#e6e6e6" else "#e6e6e6"
                # print("Current color is "+currentColor+"\nNext color is "+nextColor+"\n")
                emptyListWindow.config(bg=nextColor)
                emptyListWindow.children["emptyMessage"].config(bg = nextColor)
                emptyListWindow.children["errorOkButton"].config(highlightbackground = nextColor)
            for i in range(1,11): # Make it blink 5 times
                root.after((60*i), changeColor)
        return
    if "emptyListWindow" in globals().keys() and emptyListWindow.winfo_exists == 1: # if the error window's up and there is a valid list
        emptyListWindow.destroy()  #                                                    destory that error window
    missingImgWB = openpyxl.Workbook()
    worksheet = missingImgWB.active

    bolder = openpyxl.styles.Font(size=12,bold=True)
    a1 = worksheet.cell(1,1,"Product ID")
    a2 = worksheet.cell(1,2,"Vendor")
    a3 = worksheet.cell(1,3,"Title")
    a4 = worksheet.cell(1,4,"Link")

    a1.font = bolder
    a2.font = bolder
    a3.font = bolder
    a4.font = bolder

    for product in productList: # Append a row to the spreadhseet for each item in the product list
        worksheet.append([product[3],product[2],product[1],product[0]])

    # Adjust columns to fit values
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column # Get the column name
        for cell in col:
            try: # error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        if max_length+5 > 35:
            adjusted_width = 35
        else:
            adjusted_width = (max_length + 5)
        worksheet.column_dimensions[column].width = adjusted_width

    # Adjust top row to fit values
    # for row in worksheet.rows:
    #     max_length = 0
    #     column = col[0].column # Get the column name
    #     for cell in col:
    #         try: # Necessary to avoid error on empty cells
    #             if len(str(cell.value)) > max_length:
    #                 max_length = len(cell.value)
    #         except:
    #             pass
    #     adjusted_width = (max_length + 5)
    #     worksheet.column_dimensions[column].width = adjusted_width

    xlFileName = filedialog.asksaveasfilename(initialfile="MissingImagesSheet.xlsx")
    missingImgWB.save(filename=xlFileName)
    print("Excel sheet saved!")
    ReceivingText.config(state=NORMAL)
    ReceivingText.insert(END,"\nExcel sheet saved!")
    ReceivingText.config(state=DISABLED)
    ReceivingText.see(END)
    return
    
def makeTextResults():
    global emptyListWindow
    if not "productList" in globals().keys() or len(productList) == 0: # Make sure there's a product list with stuff in it
        if not "emptyListWindow" in globals().keys() or emptyListWindow.winfo_exists() == 0: # if there isn't already an error message up
            emptyListWindow = Toplevel(bg = "#e6e6e6")
            # emptyListWindow.wm_iconbitmap('fslogo.ico')
            if system() == "Darwin":
                pass
                # emptyListWindow.wm_iconbitmap('fslogo.icns')
            if system() == "Windows":
                emptyListWindow.wm_iconbitmap('fslogo.ico')
            emptyListWindow.title("No Products To List")
            errorText = "List of missing image products is empty.\nRun a search with results first."
            emptyMessage = Label(emptyListWindow, text=errorText, bg = "#e6e6e6", name="emptyMessage")
            errorOkButton = Button(emptyListWindow, text="OK", highlightbackground = "#e6e6e6", name="errorOkButton", command=emptyListWindow.destroy)
            emptyMessage.pack(side=TOP, anchor=N, padx=40, pady=(20,10))
            errorOkButton.pack(side=TOP, anchor=S, padx=40, pady=(0,20))
        else: #                                                              # if there IS already an error message up
            emptyListWindow.attributes('-topmost', 1) # Raise that window to the top temporarily
            emptyListWindow.attributes('-topmost', 0)
            def changeColor(): # make the window blink
                currentColor = emptyListWindow.cget("bg")
                nextColor = "#ffffff" if currentColor == "#e6e6e6" else "#e6e6e6"
                # print("Current color is "+currentColor+"\nNext color is "+nextColor+"\n")
                emptyListWindow.config(bg=nextColor)
                emptyListWindow.children["emptyMessage"].config(bg = nextColor)
                emptyListWindow.children["errorOkButton"].config(highlightbackground = nextColor)
            for i in range(1,11): # Make it blink 5 times
                root.after((60*i), changeColor)
        return
    if "emptyListWindow" in globals().keys() and emptyListWindow.winfo_exists == 1: # if the error window's up and there is a valid list
        emptyListWindow.destroy()  #                                                    destory that error window
    textResultsString = str()
    rightnow = datetime.datetime.now()
    today = str(rightnow.month)+"/"+str(rightnow.day)+"/"+str(rightnow.year)
    for i in productList:
        # [Link, Title, Vendor, ProdID]
        textResultsString += today+"\t"+nameEntry.get()+"\t"+i[3]+"\t"+i[1]+"\tMissing Image\t\t"+i[2]+"\n"
    if textResultsString[-1] == "\n":
        textResultsString = textResultsString[:-1]
    textResultsWindow = Toplevel(bg = "#e6e6e6")
    # textResultsWindow.wm_iconbitmap('fslogo.ico')
    if system() == "Darwin":
        pass
        # textResultsWindow.wm_iconbitmap('fslogo.icns')
    if system() == "Windows":
        textResultsWindow.wm_iconbitmap('fslogo.ico')
    textResultsWindow.geometry('{}x{}'.format(450, 500))
    textResultsWindow.minsize(350,400)
    textResultsWindow.title("Text Results for Copying")
    textResultsFrame = Frame(textResultsWindow, bg="#ffe6e6", bd=1, relief=SUNKEN)
    # ReceivingFrame=Frame(MainFrame, bg="#ffe6e6", bd=1, relief=SUNKEN)
    textResultsText=Text(textResultsFrame,state=DISABLED, wrap=WORD, height=15,highlightthickness=0,undo=True, font=("Helvetica",10), bd=0)
    textResultsTextScroll=Scrollbar(textResultsFrame)
    textResultsText.configure(yscrollcommand=textResultsTextScroll.set)
    textResultsTextScroll.config(command=textResultsText.yview)
    textResultsText.bind("<1>", lambda event: textResultsText.focus_set()) # Allows highlighting/copying text even when disabled
    if system() == "Darwin":
        textResultsText.bind('<Command-a>', lambda e: textResultsText.tag_add(SEL, "1.0", END))
        textResultsText.bind('<Command-A>', lambda e: textResultsText.tag_add(SEL, "1.0", END))
    if system() == "Windows":    
        textResultsText.bind('<Control-a>', lambda e: select_all("textResultsText"))
        textResultsText.bind('<Control-A>', lambda e: select_all("textResultsText"))
    textResultsFrame.pack(fill=BOTH, expand=1)
    textResultsTextScroll.pack(side=RIGHT, fill=Y)
    textResultsText.pack(side=LEFT, fill=BOTH, expand=1)
    textResultsText.config(state=NORMAL)
    textResultsText.insert(END,textResultsString)
    textResultsText.config(state=DISABLED)
    textResultsText.update_idletasks()

def reqQuery():
    global consecutiveErrors
    while True:
        try:
            query = queryQueue.get(False)
            queryThreads.append(threading.current_thread().name)
            print ("Searching '"+query+"'\nStarting...")
            updatesQueue.put(["updatemsg","Searching '"+query+"'...\n"])
            noMoreResults = False
            totalPages = 1
            pagesStep = 83
            while totalPages < 334:
                if (totalPages-1)+pagesStep >= 333:
                    searchURL = "https://flexshopper.com/search-chunk?termOrSlug="+query+"&range%5B%5D="+str(totalPages)+"&range%5B%5D=333"
                else:
                    searchURL = "https://flexshopper.com/search-chunk?termOrSlug="+query+"&range%5B%5D="+str(totalPages)+"&range%5B%5D="+str((totalPages-1)+pagesStep)
                print("Sending "+searchURL)
                try:
                    r = requests.get(searchURL)
                    consecutiveErrors = 0
                except Exception as err:
                    consecutiveErrors += 1
                    print(err)
                    errorList.append(err)
                    if consecutiveErrors >= 10:
                        earlyAbort.set()
                        return
                    continue
                print ("Received "+r.url)
                totalPages += pagesStep
                # Filter which pages have a placeholder image
                addedpages = 0
                for thispage in r.json():
                    if 'errorCode' in thispage.keys():
                        noMoreResults = not noMoreResults
                        break
                    pageQueue.put(thispage)
                    addedpages+=1
                if addedpages != 0:
                    print(threading.current_thread().name+" added "+str(addedpages)+" pages to pagesqueue.")
                if noMoreResults == True:
                    if threading.current_thread().name in queryThreads:
                        queryThreads.remove(threading.current_thread().name)
                    return
        except queue.Empty:
            if threading.current_thread().name in queryThreads:
                queryThreads.remove(threading.current_thread().name)
            return

def theMotions():
    global duplicateCount
    while AreQueuesEmpty() or AreStringsWorking(): # Run loop as long these queues aren't empty strings have tasks.
        if earlyAbort.is_set() == True:
            return
        reqQuery() # keeps going until queryQueue is empty
        addedstuff = 0
        while True:
            try:
                thispage = pageQueue.get(False)
                pageThreads.append(threading.current_thread().name)
                if 'errorCode' in thispage.keys(): # if there's an error returned for this page, declare the results ended and break the loop
                    print("errorCode found.")
                    continue
                if "product_placeholder.png" in thispage["html"]:
                    DomQueue.put(thispage["html"])
                    # print("Page with placeholder image added.")
                    addedstuff+=1
            except queue.Empty:
                if addedstuff != 0:
                    print(threading.current_thread().name+" added "+str(addedstuff)+" pages with placeholder to DomQueue.")
                # updatesQueue.put(["updatemsg","Added "+str(addedstuff)+" pages with placeholder to DomQueue.\n"])
                if threading.current_thread().name in pageThreads:
                    pageThreads.remove(threading.current_thread().name)
                break
        while True:
            try:
                mysoup = bs4.BeautifulSoup(DomQueue.get(False),"html.parser")
                DomThreads.append(threading.current_thread().name)
                for j in mysoup.select('img[src*="product_placeholder.png"]'): # get all img tags with the placeholder for a src.
                    # Gather info about the product
                    tempLink = "https://flexshopper.com"+j.parent['href']
                    if tempLink in [product[0] for product in TempProductList]: # if product link is already in the list, skip to avoid duplicates
                        duplicateCount += 1
                        continue
                    tempTitle = j.parent.parent.select('h3')[0].text.strip()
                    tempVendor = j.parent.parent.select('span[class="brandName"]')[0].text.strip()
                    TempProductList.append([tempLink,tempTitle,tempVendor])
                    needIdQueue.put([tempLink,tempTitle,tempVendor])
                    print(threading.current_thread().name+" added something to the needIdQueue.")
            except queue.Empty:
                if threading.current_thread().name in DomThreads:
                    DomThreads.remove(threading.current_thread().name)
                break
        getIdThread() # keeps going until the needIdQueue is empty

        # time.sleep(0.05) # Attempt to reduce CPU intensity

def getIdThread():
    global consecutiveErrors
    # [Link, Title, Vendor, ProdID]
    # needIdThreads.append(threading.current_thread().name)
    while True:
        try:
            prod = needIdQueue.get(False) # get next item from the queue of products.
            needIdThreads.append(threading.current_thread().name)
            try:
                s = requests.get(prod[0]) # use the link in each item to request product page
                consecutiveErrors = 0
            except Exception as err:
                consecutiveErrors += 1
                print(err)
                errorList.append(err)
                if consecutiveErrors >= 10:
                    earlyAbort.set()
                    return
                needIdQueue.put(prod)
                continue
            productsoup = bs4.BeautifulSoup(s.text,"html.parser") # parse response
            # prod.append(productsoup.select_one('h1[class="productTitle"]').text.strip()) # Get product title
            # prod.append(productsoup.select_one(".shipsFromWrapper").span.text.strip()) # Get vendor
            hiddensoup = bs4.BeautifulSoup(productsoup.select('input[class="moreChoicesModalContent"]')[0]["value"],"html.parser") # parse value of hidden button
            prod.append(hiddensoup.select('button[class*="choiceAddBtn"]')[0]["product-id"]) # get product-id from hidden button and add to list
            productList.append(prod)
            # print("\n"+prod[1]+"\n"+prod[0]+"\n"+prod[3]+"\n"+prod[2]+"\n")
            print(threading.current_thread().name+" added something to productList.")
            updatesQueue.put(["updatemsg","\n"+prod[1]+"\n"+prod[0]+"\n"+prod[3]+"\n"+prod[2]+"\n"])
        except queue.Empty:
            if threading.current_thread().name in needIdThreads:
                needIdThreads.remove(threading.current_thread().name)
            return

def justProdId():
    global consecutiveErrors
        # [Link, Title, Vendor, ProdID]
    while True:
        try:
            # http://flexshopper.com/product/5a61ec21c6a44401923f0829
            prodId = needIdQueue.get(False) # get next item from the queue of products
            prod = ["http://flexshopper.com/product/"+str(prodId)] # Get link (with ID number)
            needIdThreads.append(threading.current_thread().name)
            try:
                s = requests.get(prod[0]) # use the link in each item to request product page
            except Exception as err:
                consecutiveErrors += 1
                print(err)
                errorList.append(err)
                if consecutiveErrors >= 10:
                    earlyAbort.set()
                    return
                needIdQueue.put(prodId)
                continue
            productsoup = bs4.BeautifulSoup(s.text,"html.parser") # parse response
            if "product_placeholder.png" not in productsoup.select_one('#js-productImageFocus')["src"]:
                continue # Skip if the page doesn't have the placeholder image
            prod.append(productsoup.select_one('h1[class="productTitle"]').text.strip()) # Get product title
            prod.append(productsoup.select_one(".shipsFromWrapper").span.text.strip()) # Get vendor
            prod.append(prodId) # get product-id from hidden button and add to list
            productList.append(prod)
            # print("\n"+prod[1]+"\n"+prod[0]+"\n"+prod[3]+"\n"+prod[2]+"\n")
            print(threading.current_thread().name+" added something to productList.")
            updatesQueue.put(["updatemsg","\n"+prod[1]+"\n"+prod[0]+"\n"+prod[3]+"\n"+prod[2]+"\n"])
        except queue.Empty:
            if threading.current_thread().name in needIdThreads:
                needIdThreads.remove(threading.current_thread().name)
            break # Return and terminate the thread once the queue is empty.
    # time.sleep(0.05) # Attempt to reduce CPU intensity

def AreQueuesEmpty():
    return queryQueue.qsize() != 0 or pageQueue.qsize() != 0 or DomQueue.qsize() != 0 or needIdQueue.qsize() != 0
def AreStringsWorking():
    return len(queryThreads) !=0 or len(pageThreads) !=0 or len(DomThreads) !=0 or len(needIdThreads) !=0

def getQuery(dummy=None):
    global timerStart
    timerStart = time.time()
    global productList
    global TempProductList
    global errorList
    global duplicateCount
    duplicateCount = 0
    errorList = []
    TempProductList = []
    productList = [] # [Link, Title, Vendor, ProdID]
    print(searchCriteria.get())
    if searchCriteria.get() == "keyword":
        global queryList
        queryList = [query.strip() for query in str(searchEntry.get()).split(',')]
        print(queryList)
        [queryQueue.put(query) for query in queryList]
        print(queryQueue.qsize())
        while AreQueuesEmpty() or AreStringsWorking():
            workerThreadList = [threading.Thread(target=theMotions, name="workerThread"+str(i)) for i in range(maxThreads)]
            [thisthread.start() for thisthread in workerThreadList]
            [thisthread.join() for thisthread in workerThreadList]
            print("all threads ended")
            print(len(needIdThreads))
            print(needIdThreads)
            print(needIdQueue.qsize())
            if earlyAbort.is_set() == True:
                break
        updatesQueue.put(["updatemsg","\nFor '"+(", ").join(queryList)+"' total placeholder images: "+str(len(productList))+"\n"])
    elif searchCriteria.get() == "prodID":
        [needIdQueue.put(prodID) for prodID in findall("[a-fA-F0-9]{24}",searchEntry.get())] # get prodIDs from entry and split by their format
        print("needIdQueue size is: "+str(needIdQueue.qsize()))
        while needIdQueue.qsize() != 0 or len(needIdThreads) != 0:
            workerThreadList = [threading.Thread(target=justProdId, name="workerThread"+str(i)) for i in range(maxThreads)]
            [thisthread.start() for thisthread in workerThreadList]
            [thisthread.join() for thisthread in workerThreadList]
            print("all threads ended")
            print(len(needIdThreads))
            print(needIdThreads)
            print(needIdQueue.qsize())
            if earlyAbort.is_set() == True:
                break
    else:
        print("No Search Criteria")
        errorList.append("No Search Criteria selection.")
    print("\nProduct List: "+str(len(productList)))
    print("Temp Product List: "+str(len(TempProductList)))
    print("Duplicates rejected: "+str(duplicateCount))
    print(queryQueue.qsize(), pageQueue.qsize(), DomQueue.qsize(), needIdQueue.qsize())
    print("Time taken:")
    print(time.time() - timerStart)
    if len(errorList) != 0:
        for e in errorList:
            print(e)
    if earlyAbort.is_set() == True:
        print("Connectivity issue forced early quit. Check internet connection.")
        updatesQueue.put(["updatemsg","\nConnectivity issue forced early quit. Check internet connection.\n"])
    updatesQueue.put(["finished"])


# cycleCount = 0

def updateGUI():
    # global cycleCount
    while True:
        # cycleCount += 1
        # if cycleCount >= 10:
        #     currentStatus = """
        #     Active threads: """+str(threading.active_count())+"""
        #     queryQueue size: """+str(queryQueue.qsize())+"""
        #     pageQueue size: """+str(pageQueue.qsize())+"""
        #     DomQueue size: """+str(DomQueue.qsize())+"""
        #     needIdQueue size: """+str(needIdQueue.qsize())+"""

        #     queryThreads: """+str(len(queryThreads))+"""
        #     pageThreads: """+str(len(pageThreads))+"""
        #     DomThreads: """+str(len(DomThreads))+"""
        #     needIdThreads: """+str(len(needIdThreads))+"""
        #     """
        #     print(currentStatus)
        #     cycleCount = 0
        try:
            nextupdate = updatesQueue.get(False) # LOOP TO UPDATE THE GUI, DON'T BLOCK IF QUEUE IS EMPTY AND JUST BREAK
            if nextupdate[0] == "updatemsg":
                ReceivingText.config(state=NORMAL)
                ReceivingText.insert(END,nextupdate[1])
                ReceivingText.config(state=DISABLED)
                ReceivingText.see(END)
                ReceivingText.update_idletasks()
                if not ReceivingText.yview() == (0.0,1.0) and ReceivingTextScroll.winfo_ismapped() == False:
                    ReceivingText.pack_forget()
                    ReceivingTextScroll.pack(side=RIGHT, fill=Y)
                    ReceivingText.pack(fill=BOTH,side=LEFT,anchor=W,expand=1)
                if ReceivingText.yview() == (0.0,1.0) and ReceivingTextScroll.winfo_ismapped() == True:
                    ReceivingTextScroll.pack_forget()
                continue
            if nextupdate[0] == "noresults":
                print("Time taken:")
                print(time.time() - timerStart)
                ReceivingText.config(state=NORMAL)
                ReceivingText.insert(END,"No place holder images found.\n")
                ReceivingText.config(state=DISABLED)
                excelButton.config(state=NORMAL)
                textResultsButton.config(state=NORMAL)
                searchEntry.config(state=NORMAL)
                goButton.config(state=NORMAL)
                searchKeywordButton.config(state=NORMAL)
                searchProdIdButton.config(state=NORMAL)
                return
            if nextupdate[0] == "finished":
                consecutiveErrors = 0
                earlyAbort.clear()
                print("Time taken:")
                print(time.time() - timerStart)
                excelButton.config(state=NORMAL)
                textResultsButton.config(state=NORMAL)
                searchEntry.config(state=NORMAL)
                goButton.config(state=NORMAL)
                searchKeywordButton.config(state=NORMAL)
                searchProdIdButton.config(state=NORMAL)
                return
        except:
            break
    root.after(300,updateGUI)

##############################################################
##################        GUI BELOW       ####################
##############################################################

MainFrame = Frame(root, bg = "#e6e6e6")


nameLabel = Label(MainFrame, text="Enter name:", bg = "#e6e6e6")
nameEntry = Entry(MainFrame,width = 60, highlightbackground = "#e6e6e6", exportselection = 0)

ReceivingFrame=Frame(MainFrame, bg="#ffe6e6", bd=1, relief=SUNKEN)
ReceivingText=Text(ReceivingFrame,state=DISABLED, wrap=WORD, height=15,highlightthickness=0,undo=True, font=("Helvetica",10), bd=0)
ReceivingTextScroll=Scrollbar(ReceivingFrame)
ReceivingText.configure(yscrollcommand=ReceivingTextScroll.set)
ReceivingTextScroll.config(command=ReceivingText.yview)
ReceivingText.bind("<1>", lambda event: ReceivingText.focus_set()) # Allows highlighting/copying text even when disabled
if system() == "Darwin":
    ReceivingText.bind('<Command-a>', lambda e: ReceivingText.tag_add(SEL, "1.0", END))
    ReceivingText.bind('<Command-A>', lambda e: ReceivingText.tag_add(SEL, "1.0", END))
if system() == "Windows":    
    ReceivingText.bind('<Control-a>', lambda e: select_all("ReceivingText"))
    ReceivingText.bind('<Control-A>', lambda e: select_all("ReceivingText"))

def clearTextbox():
    ReceivingText.config(state=NORMAL)
    ReceivingText.delete("1.0", END)
    ReceivingText.config(state=DISABLED)
    ReceivingTextScroll.pack_forget()
clearButton = Button(MainFrame, text="Clear Output", highlightbackground = "#e6e6e6", command=clearTextbox)

separator1 = Frame(MainFrame,height=2,bd=1,relief=GROOVE)

def startSearch(dummy=None):
    global emptyListWindow
    searchEntry.config(state=DISABLED)
    goButton.config(state=DISABLED)
    excelButton.config(state=DISABLED)
    textResultsButton.config(state=DISABLED)
    searchKeywordButton.config(state=DISABLED)
    searchProdIdButton.config(state=DISABLED)

    if "emptyListWindow" in globals().keys() and emptyListWindow.winfo_exists() == 1:
        emptyListWindow.destroy()
    searchThread = threading.Thread(target=getQuery, name="SearchProcessThread",args=(), daemon=True)
    searchThread.start()
    updateGUI()

def mytest():
    print(searchCriteria.get())
    
searchCriteriaFrame = Frame(MainFrame, bg = "#e6e6e6")
searchKeywordButton = Radiobutton(searchCriteriaFrame, bg="#e6e6e6", text="Serch by keyword", state=NORMAL, \
    variable=searchCriteria, value="keyword", command="searchCriteria.set()")
searchProdIdButton = Radiobutton(searchCriteriaFrame, bg="#e6e6e6", text="Search by Product ID", state=NORMAL, \
    variable=searchCriteria, value="prodID", command="searchCriteria.set()")
searchKeywordButton.select()
searchProdIdButton.deselect()
searchLabel = Label(MainFrame, text="Separate keywords with commas, Product IDs by non-alphanumeric characters:", bg = "#e6e6e6")
searchEntry = Entry(MainFrame,width = 60, highlightbackground = "#e6e6e6", exportselection = 0)
searchEntry.bind('<Return>', startSearch)
if system() == "Darwin":
    searchEntry.bind('<Command-a>', lambda e: searchEntry.selection_range(0, END))
    searchEntry.bind('<Command-A>', lambda e: searchEntry.selection_range(0, END))
if system() == "Windows":    
    searchEntry.bind('<Control-a>', lambda e: select_all("searchEntry"))
    searchEntry.bind('<Control-A>', lambda e: select_all("searchEntry"))
ButtonFrame1 = Frame(MainFrame, bg = "#e6e6e6")
goButton = Button(ButtonFrame1, text="Start Search", highlightbackground = "#e6e6e6", command=startSearch)
excelButton = Button(ButtonFrame1, text="Make Excel File", highlightbackground = "#e6e6e6", command=makeExcelDoc)
textResultsButton = Button(ButtonFrame1, text="Copiable Results", highlightbackground = "#e6e6e6", command=makeTextResults)
testButton = Button(ButtonFrame1, text="Test", highlightbackground = "#e6e6e6", command=mytest)


MainFrame.pack(fill = BOTH, expand = 1)

##### FROM THE BOTTOM #####
ButtonFrame1.pack(side=BOTTOM, padx=10,pady=10, anchor=S, fill=X, expand=1) # Mainframe
goButton.pack(side=RIGHT, padx=0,pady=0, anchor=E)
excelButton.pack(side=LEFT, padx=0,pady=0, anchor=W)
textResultsButton.pack(side=LEFT, padx=10,pady=0, anchor=W)
searchEntry.pack(side=BOTTOM,padx=10,pady=(5,10), anchor=SW) # Mainframe
searchLabel.pack(side=BOTTOM,padx=10, pady=(10,5), anchor=SW) # Mainframe
searchCriteriaFrame.pack(side=BOTTOM, expand=1, fill=X, anchor=SW, padx=10, pady=(10,5))
searchKeywordButton.pack(side=LEFT, anchor=W)
searchProdIdButton.pack(side=LEFT, anchor=W)
# testButton.pack(side=LEFT, anchor=W)

separator1.pack(side=BOTTOM, padx=60, pady=10, fill=X)

##### FROM THE TOP #####
nameLabel.pack(side=TOP, padx=27, pady=(10,5), anchor=NW)
nameEntry.pack(side=TOP, padx=30, pady=(0,0), anchor=NW)
ReceivingFrame.pack(side=TOP, fill=BOTH, expand=1, padx=30, pady=(20,0), anchor=N)
clearButton.pack(side=TOP, anchor=W, padx=30,pady=(10,10))
# ReceivingTextScroll.pack(side=RIGHT, fill=Y)
ReceivingText.pack(fill=BOTH,side=LEFT,anchor=W,expand=1)

print("Active count: "+str(threading.active_count()))
root.mainloop()

